from openpyxl import *
from tkinter import *
import tkinter as tk
from tkinter import tk, Canvas
from tkinter.tk import *
from tkinter import messagebox as mb
import pandas as pd
from PIL import ImageTk, Image
import pycountry
from tkcalendar import DateEntry
from datetime import date

def excel(sheet): 
    # at particular location 
    sheet.cell(row=1, column=1).value = "Timestamp"
    sheet.cell(row=1, column=2).value = "Active Power"
    sheet.cell(row=1, column=3).value = "Wind Speed"
    sheet.cell(row=1, column=4).value = "Wind Direction"
    sheet.cell(row=1, column=5).value = "Country"

def insert(sheet, active_pow, wind_speed ,wind_dir, country, datetime):
    current_row = sheet.max_row 
    current_column = sheet.max_column
    sheet.cell(row = current_row + 1, column = 1).value = datetime
    sheet.cell(row = current_row + 1, column = 2).value = active_pow
    sheet.cell(row = current_row + 1, column = 3).value = wind_speed
    sheet.cell(row = current_row + 1, column = 4).value = wind_dir
    sheet.cell(row = current_row + 1, column = 5).value = country

def dn_clicked():
    mb.showinfo('Thank you!')
    window.destroy()

def btn_clicked():
    d = dt.get()
    t = time.get()
    datetime = str(d + " " + t)
    active_pow = txt_ap.get()
    wind_speed = txt_ws.get()
    wind_dir = txt_wd.get()
    country = txt_country.get()
    if (date and time and active_pow and wind_speed and wind_dir and len(t) == 8):
        #the user entered data in the mandatory entry: proceed to next step
        mb.showinfo('Wind Turbine Data Submission','Press OK to start calculating your data....\n')
        #model call
        value = 0
        #displaying predicted value
        ap_out = tk.Message(window, text = str(value), background = "white")
        ap_out.place(x = 240, y = 360)
        insert(sheet, active_pow, wind_speed, wind_dir, country, datetime)
        window.mainloop()
        #saving the dataframe 
        wb.save('wind_turbine.xlsx')
        data_xls = pd.read_excel('wind_turbine.xlsx', index_col = None)
        data_xls.to_csv('wind_turbine.csv', encoding = 'utf-8', index = False)
    else:
        #the mandatory field is empty
        mb.showwarning('Wind Turbine Data Submission','Some fields are empty!')

def cncl_clicked():
    mb.showinfo('Wind Turbine Data Submission','Your feedback will be cancelled.')
    window.destroy()

#load workbook
def load():
    wb = load_workbook("wind_turbine.xlsx")  
    sheet = wb.active
    return wb, sheet

#window
def create_window():
    #main window
    window = Tk()
    window.title("Wind Turbine Data")
    window.geometry('1200x450')
    window.iconbitmap(default = "logo.ico")
    window.config(background = "white")
    #window style
    tk.Style().configure('Wild.TRadiobutton', background = "white")
    tk.Style().configure('Wild.TLabel', background = "light cyan", padding = 6, relief="groove")
    return window

def create_labels(window, path):
    #separator
    Canvas(window).create_line(240, 50, 240, 350)
    #heading
    lbl = tk.Label(window, text = "WIND TURBINE DATA INPUT", font = ("BOLD", 14), background = "white")
    lbl.place(x = 50, y = 15)
    lbl = tk.Label(window, text = "Time Series Trend", font = ("BOLD", 14), background = "white")
    lbl.place(x = 700, y = 15)
    #active power
    txt = tk.Label(window, text = "Active Power(kWh):", background = "white")
    txt.place(x = 50, y = 50)
    txt_ap = tk.Entry(window, width = 20)
    txt_ap.place(x = 50, y = 70)
    #wind speed
    txt = tk.Label(window, text = "Wind Speed(kmph):", background = "white")
    txt.place(x = 50, y = 100)
    txt_ws = tk.Entry(window, width = 20)
    txt_ws.place(x = 50, y = 120)
    #wind direction
    txt = tk.Label(window, text = "Wind Direction(deg):", background = "white")
    txt.place(x = 50, y = 150)
    txt_wd = tk.Entry(window, width = 20)
    txt_wd.place(x = 50, y = 170)
    #country
    txt = tk.Label(window, text = "Country:", background = "white")
    txt.place(x = 50, y = 200)
    txt_country = Combobox(window, values = ["", "Europe", "North America", "South America", "Africa", "Asia", "Antarctica", "Oceania"])
    txt_country.place(x = 50, y = 220)
    txt_country.current(0)
    #date
    txt = tk.Label(window, text = "Date(MM/DD/YYYY):", background = "white")
    txt.place(x = 50, y = 250)
    dt = DateEntry(window, locale = 'en_US', date_pattern='MM/dd/yyyy')
    dt.place(x = 50, y = 270)
    #time
    txt = tk.Label(window, text = "Time(hr:min:sec):", background = "white")
    txt.place(x = 180, y = 250)
    time = tk.Entry(window, width = 8)
    time.place(x = 180, y = 270)
    #predicted output
    txt = tk.Label(window, text = "Active Power(predicted):", background = "white")
    txt.place(x = 50, y = 360)
    #time series display
    img = ImageTk.PhotoImage(Image.open(path).resize((750, 200), Image.ANTIALIAS))
    ts_pic = tk.Label(window, image = img)
    ts_pic.place(x = 400, y = 50)
    return txt_ap, txt_ws, txt_wd, txt_country, dt, time

#submit or cancel
def create_sub_can():
    tk.Button(window, text = "Done", style = "Wild.TLabel", command = dn_clicked).place(x = 50, y = 400)
    tk.Button(window, text = "Submit", style = "Wild.TLabel", command = btn_clicked).place(x = 50, y = 300)
    tk.Button(window, text = "Cancel", style = "Wild.TLabel", command = cncl_clicked).place(x = 140, y = 400)

#sequence
wb, sheet = load()
excel(sheet)
window = create_window()
path = 'ws_wd_ts.png' #time series image path
txt_ap, txt_ws, txt_wd, txt_country, dt, time = create_labels(window, path)
create_sub_can()
window.mainloop()









